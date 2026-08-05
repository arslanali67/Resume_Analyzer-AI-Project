from pathlib import Path
from openpyxl.formatting.rule import CellIsRule
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment,
)
HEADER_FILL = PatternFill(
    fill_type="solid",
    start_color="1F4E78",
    end_color="1F4E78",
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF",
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CENTER = Alignment(
    horizontal="center",
    vertical="center",
)

GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="C6EFCE",
    end_color="C6EFCE",
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    start_color="FFEB9C",
    end_color="FFEB9C",
)

RED_FILL = PatternFill(
    fill_type="solid",
    start_color="FFC7CE",
    end_color="FFC7CE",
)


def auto_size_sheet(sheet):
    """
    Automatically adjust column widths.
    """

    for column_cells in sheet.columns:

        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:

            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

        sheet.column_dimensions[column_letter].width = min(
            max_length + 3,
            60,
        )


def style_header(sheet):
    """
    Apply professional styling to the header row.
    """

    for cell in sheet[1]:

        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER

def export_to_excel(results):
    """
    Export ranked evaluation results to Excel.
    """

    # Create output folder
    output_folder = Path("output")
    output_folder.mkdir(exist_ok=True)

    output_file = output_folder / "evaluation_results.xlsx"

    workbook = Workbook()

    # =====================================================
    # Sheet 1 : Candidate Ranking
    # =====================================================

    ranking_sheet = workbook.active
    ranking_sheet.title = "Candidate Ranking"

    ranking_headers = [
        "Rank",
        "Candidate",
        "Resume",
        "Score",
        "Recommendation",
        "Experience (Years)",
        "Current Role",
        "Email",
        "Phone",
        "Location",
    ]

    for col, header in enumerate(ranking_headers, start=1):

        ranking_sheet.cell(
            row=1,
            column=col,
        ).value = header

    style_header(ranking_sheet)

    # =====================================================
    # Sheet 2 : Detailed Evaluation
    # =====================================================

    detail_sheet = workbook.create_sheet(
        title="Detailed Evaluation"
    )

    detail_headers = [
        "Candidate",
        "Resume",
        "Score",
        "Matching Skills",
        "Missing Skills",
        "Strengths",
        "Weaknesses",
        "Experience Summary",
        "Recommendation",
    ]

    for col, header in enumerate(detail_headers, start=1):

        detail_sheet.cell(
            row=1,
            column=col,
        ).value = header

    style_header(detail_sheet)

    # =====================================================
    # Sheet 3 : Candidate Metadata
    # =====================================================

    metadata_sheet = workbook.create_sheet(
        title="Candidate Metadata"
    )

    metadata_headers = [
        "Candidate",
        "Resume",
        "Email",
        "Phone",
        "Education",
        "Experience (Years)",
        "Current Role",
        "Location",
    ]

    for col, header in enumerate(metadata_headers, start=1):

        metadata_sheet.cell(
            row=1,
            column=col,
        ).value = header

    style_header(metadata_sheet)

    # =====================================================
    # Write Data
    # =====================================================

    for row, result in enumerate(results, start=2):

        evaluation = result["evaluation"]
        metadata = result["metadata"]
        
        # --------------------------
        # Candidate Ranking
        # --------------------------

        ranking_sheet.cell(row=row, column=1).value = row - 1
        ranking_sheet.cell(row=row, column=2).value = metadata.candidate_name
        ranking_sheet.cell(row=row, column=3).value = result["filename"]
        ranking_sheet.cell(row=row, column=4).value = evaluation.match_score
        ranking_sheet.cell(row=row, column=5).value = evaluation.hiring_recommendation
        ranking_sheet.cell(row=row, column=6).value = metadata.experience_years
        ranking_sheet.cell(row=row, column=7).value = metadata.current_role
        ranking_sheet.cell(row=row, column=8).value = metadata.email
        ranking_sheet.cell(row=row, column=9).value = metadata.phone
        ranking_sheet.cell(row=row, column=10).value = metadata.location

        # --------------------------
        # Detailed Evaluation
        # --------------------------

        detail_sheet.cell(row=row, column=1).value = metadata.candidate_name
        detail_sheet.cell(row=row, column=2).value = result["filename"]
        detail_sheet.cell(row=row, column=3).value = evaluation.match_score
        detail_sheet.cell(row=row, column=4).value = ", ".join(
            evaluation.matching_skills
        )
        detail_sheet.cell(row=row, column=5).value = ", ".join(
            evaluation.missing_skills
        )
        detail_sheet.cell(row=row, column=6).value = ", ".join(
            evaluation.strengths
        )
        detail_sheet.cell(row=row, column=7).value = ", ".join(
            evaluation.weaknesses
        )
        detail_sheet.cell(row=row, column=8).value = (
            evaluation.experience_summary
        )
        detail_sheet.cell(row=row, column=9).value = (
            evaluation.hiring_recommendation
        )

        # --------------------------
        # Candidate Metadata
        # --------------------------

        metadata_sheet.cell(row=row, column=1).value = metadata.candidate_name
        metadata_sheet.cell(row=row, column=2).value = result["filename"]
        metadata_sheet.cell(row=row, column=3).value = metadata.email
        metadata_sheet.cell(row=row, column=4).value = metadata.phone
        metadata_sheet.cell(row=row, column=5).value = metadata.education
        metadata_sheet.cell(row=row, column=6).value = metadata.experience_years
        metadata_sheet.cell(row=row, column=7).value = metadata.current_role
        metadata_sheet.cell(row=row, column=8).value = metadata.location

    # =====================================================
    # Auto Size Columns
    # =====================================================

    auto_size_sheet(ranking_sheet)
    auto_size_sheet(detail_sheet)
    auto_size_sheet(metadata_sheet)

    # =====================================================
    # Freeze Header Row
    # =====================================================

    ranking_sheet.freeze_panes = "A2"
    detail_sheet.freeze_panes = "A2"
    metadata_sheet.freeze_panes = "A2"

    # =====================================================
    # Enable Filters
    # =====================================================

    ranking_sheet.auto_filter.ref = ranking_sheet.dimensions
    detail_sheet.auto_filter.ref = detail_sheet.dimensions
    metadata_sheet.auto_filter.ref = metadata_sheet.dimensions

    # =====================================================
    # Conditional Formatting (Ranking Sheet)
    # =====================================================

    ranking_sheet.conditional_formatting.add(
        "D2:D1000",
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["80"],
            fill=GREEN_FILL,
        ),
    )

    ranking_sheet.conditional_formatting.add(
        "D2:D1000",
        CellIsRule(
            operator="between",
            formula=["50", "79"],
            fill=YELLOW_FILL,
        ),
    )

    ranking_sheet.conditional_formatting.add(
        "D2:D1000",
        CellIsRule(
            operator="lessThan",
            formula=["50"],
            fill=RED_FILL,
        ),
    )

    # =====================================================
    # Conditional Formatting (Detailed Evaluation Sheet)
    # =====================================================

    detail_sheet.conditional_formatting.add(
        "C2:C1000",
        CellIsRule(
            operator="greaterThanOrEqual",
            formula=["80"],
            fill=GREEN_FILL,
        ),
    )

    detail_sheet.conditional_formatting.add(
        "C2:C1000",
        CellIsRule(
            operator="between",
            formula=["50", "79"],
            fill=YELLOW_FILL,
        ),
    )

    detail_sheet.conditional_formatting.add(
        "C2:C1000",
        CellIsRule(
            operator="lessThan",
            formula=["50"],
            fill=RED_FILL,
        ),
    )

# =====================================================
# Save Workbook
# =====================================================


    # Save workbook
    workbook.save(output_file)

    print(f"\nExcel report saved to: {output_file}")

